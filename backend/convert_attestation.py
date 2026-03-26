#!/usr/bin/env python3
"""
One-time script: parse attestation Excel and append entries to benchmark_data.json.

Usage:
    python3 convert_attestation.py /path/to/attestation.xlsx
"""
import json
import re
import sys
from pathlib import Path

import openpyxl


def parse_questions(xlsx_path: str) -> list[dict]:
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    questions: list[dict] = []
    current: dict | None = None

    for r in range(1, ws.max_row + 1):
        a = ws.cell(r, 1).value
        b = ws.cell(r, 2).value
        c = ws.cell(r, 3).value

        # Skip header/meta rows
        if isinstance(a, str) and ("Попытка" in a or "№" in a):
            continue

        if isinstance(a, (int, float)) and a > 0 and b:
            # New question
            if current:
                questions.append(current)
            current = {
                "num": int(a),
                "question": str(b).strip(),
                "correct_answer": str(c).strip() if c else "",
                "options": [str(c).strip()] if c else [],
            }
        elif current and c and str(c).strip():
            current["options"].append(str(c).strip())

    if current:
        questions.append(current)

    return questions


def extract_facts(correct_answer: str) -> list[str]:
    """Extract key factual substrings from the correct answer."""
    facts: list[str] = []
    # Extract numbers with units (e.g. "40гр", "140гр", "23 секунд")
    for m in re.finditer(r'\d+[\s]?(?:гр|мл|секунд|минут|%|шт|раз)', correct_answer):
        facts.append(m.group().replace(" ", ""))
    # If no numeric facts found, take first meaningful word (5+ chars)
    if not facts:
        for word in correct_answer.split():
            clean = word.strip(".,;:!?()\"'")
            if len(clean) >= 5:
                facts.append(clean.lower())
                break
    return facts


def to_benchmark_entries(questions: list[dict]) -> list[dict]:
    entries = []
    for i, q in enumerate(questions, 1):
        wrong = [opt for opt in q["options"] if opt != q["correct_answer"]]
        entry = {
            "id": f"attest_{i:03d}",
            "category": "attestation",
            "question": q["question"],
            "expected_facts": extract_facts(q["correct_answer"]),
            "expected_sources": [],
            "expect_images": False,
            "expect_found": True,
            "correct_answer": q["correct_answer"],
            "wrong_answers": wrong,
        }
        entries.append(entry)
    return entries


def main():
    if len(sys.argv) < 2:
        print("Usage: python3 convert_attestation.py <xlsx_path>")
        sys.exit(1)

    xlsx_path = sys.argv[1]
    questions = parse_questions(xlsx_path)
    print(f"Parsed {len(questions)} questions from Excel.")

    entries = to_benchmark_entries(questions)

    # Load existing benchmark data
    data_path = Path(__file__).parent / "benchmark_data.json"
    with open(data_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    # Remove existing attestation entries (idempotent)
    data["questions"] = [q for q in data["questions"] if q.get("category") != "attestation"]

    # Append new attestation entries
    data["questions"].extend(entries)

    with open(data_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
        f.write("\n")

    print(f"Added {len(entries)} attestation entries to {data_path}")
    print(f"Total questions: {len(data['questions'])}")

    # Preview
    for e in entries[:3]:
        print(f"\n  {e['id']}: {e['question'][:60]}...")
        print(f"    correct: {e['correct_answer'][:60]}")
        print(f"    facts: {e['expected_facts']}")
        print(f"    wrong: {len(e['wrong_answers'])} options")


if __name__ == "__main__":
    main()
